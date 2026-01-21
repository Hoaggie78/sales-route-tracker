import openpyxl
from datetime import datetime
from typing import List, Dict
import re


def parse_customer_cell(cell_value: str) -> Dict[str, str]:
    """Parse customer information from cell value"""
    if not cell_value:
        return None
    
    lines = cell_value.strip().split('\n')
    if len(lines) < 3:
        return None
    
    name = lines[0].strip()
    address = lines[1].strip()
    
    # Extract account number
    account_match = re.search(r'Acct:\s*(\w+)', lines[2])
    account_number = account_match.group(1) if account_match else ""
    
    return {
        "name": name,
        "address": address,
        "account_number": account_number
    }


def parse_date_from_header(date_str: str) -> datetime:
    """Parse date from header like 'MONDAY\n01/19/2026'"""
    if not date_str:
        return None
    
    lines = date_str.split('\n')
    if len(lines) < 2:
        return None
    
    day_of_week = lines[0].strip()
    date_part = lines[1].strip()
    
    try:
        parsed_date = datetime.strptime(date_part, "%m/%d/%Y")
        return parsed_date.date(), day_of_week
    except:
        return None, None


def parse_excel_route_plan(file_path: str) -> List[Dict]:
    """Parse the entire Excel route plan and return structured data"""
    wb = openpyxl.load_workbook(file_path)
    
    if '4-Week Route Plan' not in wb.sheetnames:
        raise ValueError(f"Sheet '4-Week Route Plan' not found. Available sheets: {wb.sheetnames}")
        
    ws = wb['4-Week Route Plan']
    
    customers = []
    
    # Define week starting rows
    week_configs = [
        {"week_number": 1, "week_label": "WEEK 1", "start_row": 5},
        {"week_number": 2, "week_label": "WEEK 2", "start_row": 22},
        {"week_number": 3, "week_label": "WEEK 3", "start_row": 39},
        {"week_number": 4, "week_label": "WEEK 4", "start_row": 56},
    ]
    
    for week_config in week_configs:
        week_number = week_config["week_number"]
        week_label = week_config["week_label"]
        header_row = week_config["start_row"]
        location_row = header_row + 1
        
        # Parse dates and days from header row
        dates_info = []
        for col_idx in range(2, 7):  # Columns B through F (Monday-Friday)
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value:
                date_obj, day_of_week = parse_date_from_header(cell_value)
                dates_info.append({
                    "date": date_obj,
                    "day_of_week": day_of_week,
                    "col_idx": col_idx
                })
        
        # Parse locations
        locations = []
        for col_idx in range(2, 7):
            location = ws.cell(row=location_row, column=col_idx).value
            locations.append(location if location else "")
        
        # Parse customers (rows with stop numbers 1-10)
        for stop_number in range(1, 11):
            customer_row = header_row + 2 + stop_number  # +2 for location and header rows
            
            # Check if this row has the stop number
            stop_cell = ws.cell(row=customer_row, column=1).value
            if stop_cell != stop_number:
                continue
            
            # Parse each day's customer
            for day_idx, date_info in enumerate(dates_info):
                col_idx = date_info["col_idx"]
                cell_value = ws.cell(row=customer_row, column=col_idx).value
                
                if not cell_value:
                    continue
                
                customer_data = parse_customer_cell(cell_value)
                if not customer_data:
                    continue
                
                customers.append({
                    **customer_data,
                    "week_number": week_number,
                    "week_label": week_label,
                    "day_of_week": date_info["day_of_week"],
                    "date": date_info["date"],
                    "location": locations[day_idx],
                    "stop_number": stop_number
                })
    
    return customers


def export_tracking_data(customers_with_visits: List[Dict], output_path: str):
    """Export tracking data back to a new Excel sheet"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visit Tracking"
    
    # Headers
    headers = [
        "Week", "Day", "Date", "Location", "Stop #",
        "Customer Name", "Account #", "Address",
        "Status", "Visited At", "Sales Amount",
        "Notes", "Follow-up Required", "Follow-up Date"
    ]
    ws.append(headers)
    
    # Data rows
    for customer in customers_with_visits:
        visit = customer.get("latest_visit", {})
        ws.append([
            customer["week_label"],
            customer["day_of_week"],
            customer["date"].strftime("%m/%d/%Y") if customer["date"] else "",
            customer["location"],
            customer["stop_number"],
            customer["name"],
            customer["account_number"],
            customer["address"],
            visit.get("status", "not_visited"),
            visit.get("visited_at", ""),
            visit.get("sales_amount", 0.0),
            visit.get("notes", ""),
            visit.get("follow_up_required", False),
            visit.get("follow_up_date", "")
        ])
    
    wb.save(output_path)
